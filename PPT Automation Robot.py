from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
# import csv
import time
import os
# -------------------------------------------------------------------------

# initialize the object
Robot = webdriver.Chrome()
Robot.maximize_window()
Robot.get("Web App Link")
hold = WebDriverWait(Robot,30)


def spinner() :
    WebDriverWait(Robot, 20).until(EC.presence_of_element_located((By.CLASS_NAME, "CLASS_NAME")))

# Account Lines
account_lines_k = 0

# Account Passed Or Not
account_passed = True

# Global Var to be called when i need at the fx
city = ""


# Extra Mile
def invisibility_overlay():
    """Wait till overlay disappear when navigate"""
    return hold.until(EC.invisibility_of_element_located((By.ID, "ID" )))


def scrolling():
    """Control Scroll for all"""
    Robot.execute_script("window.scrollTo(0,550);")



def login_page(username,password) :
    """Login FX"""

    # Check Point To Move.
    hold.until(EC.visibility_of_element_located((By.XPATH,"Footer")))
    # Enter UserName
    hold.until(EC.visibility_of_element_located((By.ID,"username"))).send_keys(username)
    # Enter Password
    hold.until(EC.visibility_of_element_located((By.ID, "password"))).send_keys(password)
    # Press Ok To log in
    hold.until(EC.element_to_be_clickable((By.CLASS_NAME, "Button Standard"))).click()

    hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Tab"))).click()



def full_cycle(account_num):
    global account_passed
    account_container = []
    """Start From Source Page And Run Full Cycle"""
    try :
        def page_source(acc_num):
            """Enter account number to pass source page"""
            global city

            # Press Search to view
            hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search"))).click()
            # Check New Page Elements visibility
            hold.until(EC.visibility_of_element_located((By.XPATH, "footer")))

            # select from Drop Menu
            drop_menu = hold.until(EC.element_to_be_clickable((By.CLASS_NAME, "Select Drop Menu")))
            choice = Select(drop_menu)
            choice.select_by_index(0)

            # Enter The Account Number
            hold.until(EC.visibility_of_element_located((By.ID, "Id Code"))).send_keys(acc_num)

            # Press Search.
            hold.until(EC.element_to_be_clickable((By.ID, "Search Button"))).click()

            # City
            ccity = hold.until(EC.visibility_of_all_elements_located((By.XPATH, "Path")))[3].text

            # Press to enter
            hold.until(EC.element_to_be_clickable((By.CLASS_NAME, "CLASS_NAME"))).click()

        # Run Fx of source page
        page_source(account_num)
        invisibility_overlay()


        def home_page():
            """Fx Full Loop"""
            global account_lines_k

            # Check Point To Move Using Footer
            hold.until(EC.visibility_of_element_located((By.XPATH, "footer")))
            # Check Point To Move Using Pro Section
            hold.until(EC.visibility_of_element_located((By.XPATH, "//div[@class='class name']//div[@id='id name']//span")))


            def re_fetch() :
                """Re-Fetching Home Page For Each Call"""
                # Check Point To Move / Validation Before Go
                return hold.until(EC.visibility_of_all_elements_located((By.XPATH, "//table[@id='id name']//tbody//tr")))
            re_fetch()


            # Static Variable/Data.........
            name = hold.until(EC.visibility_of_all_elements_located((By.CLASS_NAME, "name")))[0].text
            p_class = hold.until(EC.visibility_of_all_elements_located((By.CLASS_NAME, "name")))[8].text[0:8]

            bc = hold.until(EC.visibility_of_element_located((By.XPATH, "//*[contains(text(),'Postpaid BC')]"))).text[-1]
            dials_k = len(re_fetch())


            # Global access
            account_lines_k = dials_k

            open_amount = hold.until(EC.visibility_of_all_elements_located((By.CLASS_NAME, "class name")))[3].text

            def method_lang() :

                """This Method To Get rest data"""

                hold.until(EC.element_to_be_clickable((By.ID,"ID Name"))).click()
                invisibility_overlay()
                scrolling()
                p_method = hold.until(EC.visibility_of_all_elements_located((By.XPATH,"//table[@id='ID name']//tbody//tr[@class='DATblRowA']//td")))[2].text
                time.sleep(0.5)
                hold.until(EC.element_to_be_clickable((By.ID,"Id name"))).click()
                invisibility_overlay()
                scrolling()
                language = hold.until(EC.visibility_of_all_elements_located((By.CLASS_NAME,"class name")))[4].text
                time.sleep(0.5)

                # Turn Back to Home Page
                hold.until(EC.element_to_be_clickable((By.ID,"Id name"))).click()
                invisibility_overlay()
                re_fetch()

                return p_method,language

            acc_payment_method , acc_language = method_lang()


            # call global variable
            global city

            # Loop Phase As Of Dynamic Data
            for one_dial in range(len(re_fetch())) :
                dial_row = []
                # re_fetch() the home Page Each Cycle
                re_fetch()

                # str concept to avoid end zero truncate
                dial_row.append(f"Account {account_num}")

                cst_dial = re_fetch()[one_dial].find_elements(By.TAG_NAME, "td")[1].text
                dial_row.append(cst_dial)

                rate_plan = re_fetch()[one_dial].find_elements(By.TAG_NAME, "td")[3].text
                dial_row.append(rate_plan)

                activation_date = re_fetch()[one_dial].find_elements(By.TAG_NAME, "td")[5].text
                dial_row.append(activation_date)


                # Scroll Using JS Execute
                scrolling()


                dial_row.append(bc)
                dial_row.append(p_class)
                dial_row.append(dials_k)
                dial_row.append(name)
                dial_row.append(acc_payment_method)
                dial_row.append(acc_language)
                dial_row.append(city)


                # move_to_inner_page() :
                WebDriverWait(re_fetch()[one_dial],20).until(EC.element_to_be_clickable((By.XPATH,"XPATH"))).click()

                spinner()

                def ser_num():
                    """Check Point using sr num"""
                    return hold.until(EC.visibility_of_all_elements_located((By.XPATH, "XPATH")))[1].text
                ser_num()

                # Start To Extracting Inner Page Data
                status_elements = hold.until(EC.visibility_of_all_elements_located((By.XPATH, "//span[@class='class name']")))
                status = status_elements[0].text
                dial_row.append(status)
                status_reason = status_elements[1].text
                dial_row.append(status_reason)
                dial_row.append(f"Serial {ser_num()}")
                dial_row.append(open_amount)


                # Return Back TO Previous Page
                spinner()
                time.sleep(0.05)
                step_back = hold.until(EC.element_to_be_clickable((By.XPATH,'XPATH']//a")))
                step_back.click()

                spinner()
                re_fetch()
                time.sleep(0.15)

                # Append The Full Data
                account_container.append(dial_row)

        # FX Of Home Page
        home_page()


    except Exception as e:
        Robot.refresh()
        # Global Access
        account_passed = False
        Robot.save_screenshot(f"error_screenshot{account_num}.png")
        
        # In Case Of Error Occur
        print(f"Issued Account # {account_num}")
        print(f"Error : {e}")
        # Set Refresh for Page Here ?
        print(f"Failed : {account_num}")
        time.sleep(1.5)

        # invisibility_overlay()
        point_zero = hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search")))
        point_zero.click()

    else:
        # if there is no error occur
        account_passed = True
        print(f"Account Number Succeeded : {account_num}")


    finally:
        # def re_fetch():
        # """Re-Fetching Home Page For Each Call"""
        # # Check Point To Move / Validation For All Account Dials Before Go
        # return hold.until(
        # EC.visibility_of_all_elements_located((By.XPATH, "XPATH")))
        # re_fetch()
        # spinner()


        time.sleep(1)
        invisibility_overlay()
        # Perform regardless the try/except result
        point_zero = hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search")))
        point_zero.click()

        # Check Point using footer
        hold.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='id name']//p")))


        time.sleep(0.5)

    return account_container



# Start Engine Sequence
login_page("user name","password")

# File Handling || Source Phase--1
source_file_path = r"Path"
source_wb = excel.load_workbook(source_file_path)
source_sheet = source_wb["Sheet1"]

# File Handling || Destination Phase--2
destination_file_path = r"Path"
destination_wb = excel.load_workbook(destination_file_path)
destination_sheet = destination_wb["Sheet2"]



# Loop Over Source
for axis in range( 2 , source_sheet.max_row+1 ) :
    print(f"[{axis}] Account || in progress....")
    # Take from float and str format as the ( zero ) value maybe truncated if not str
    account_number = str(source_sheet.cell(axis, 1).value)

    # List Value
    result_full_account_data = full_cycle(account_number)

    account_status = source_sheet.cell(axis, 3)

    if account_passed == True:
        account_status.value = "Done"
        source_sheet.cell(axis, 4).value = account_lines_k


    elif account_passed == False:
        account_status.value = "Not Done"
        source_sheet.cell(axis, 4).value = "Null"


    source_wb.save(source_file_path)


    # Append the List of lists
    for each_row in result_full_account_data :
        destination_sheet.append(each_row)
    destination_wb.save(destination_file_path)

print("Exting.....")
time.sleep(2)
Robot.quit()



